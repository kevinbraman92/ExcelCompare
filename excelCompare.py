import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def compare_excel_files(firstFile, secondFile, outputFileName):
    inputFolder = 'ExcelFiles'
    output_folder = 'ExcelOutput'
    if not (os.path.isdir(inputFolder) and os.path.isdir(output_folder)):
        print("Creating 'ExcelFiles' & 'ExcelOutput' folders.")
        print("Please run this file again to compare the files wihin the 'ExcelFiles' folder.")
        os.makedirs(inputFolder, exist_ok=True)
        os.makedirs(output_folder, exist_ok=True)
        print("\nExiting program...")
        sys.exit()
    if not os.listdir(inputFolder):
        print("ExcelFiles folder is empty! Please put your two Excel files in there before running this program!")
        print("\nExiting program...")
        sys.exit()
    
    print("The following prompts will ask you to enter a file name. You do not need to include the extension.")
    while True:
        try:
            firstFile = input('\nPlease enter the name of the first file:  ') + '.xlsx'
            dataFrameFirstFile = pd.read_excel(os.path.join(inputFolder, firstFile), sheet_name=0, engine='openpyxl')
            break
        except FileNotFoundError:
            print(f"Error! The file '{firstFile}' could not be found in the '{inputFolder}' folder.")
            print("Please try again.")
    while True:
        try:
            secondFile = input('\nPlease enter the name of the second file: ') + '.xlsx'
            dataFrameSecondFile = pd.read_excel(os.path.join(inputFolder, secondFile), sheet_name=0, engine='openpyxl')
            break
        except FileNotFoundError:
            print(f"Error! The file '{secondFile}' could not be found in the '{inputFolder}' folder.")
            print("Please try again.")
    
    print(f"\nComparing {firstFile} and {secondFile}...")

    reshapedRows = max(dataFrameFirstFile.shape[0], dataFrameSecondFile.shape[0])
    reshapedColumns = max(dataFrameFirstFile.shape[1], dataFrameSecondFile.shape[1])
    dataFrameFirstFile = dataFrameFirstFile.reindex(index=range(reshapedRows), columns=dataFrameFirstFile.columns.tolist() + [None]*(reshapedColumns - dataFrameFirstFile.shape[1]))
    dataFrameSecondFile = dataFrameSecondFile.reindex(index=range(reshapedRows), columns=dataFrameSecondFile.columns.tolist() + [None]*(reshapedColumns - dataFrameSecondFile.shape[1]))

    outputFilePath = os.path.join(output_folder, f"Compared_{firstFile.replace('.xlsx', '')}_vs_{secondFile}")
    with pd.ExcelWriter(outputFilePath, engine='openpyxl') as writer:
        dataFrameFirstFile.to_excel(writer, sheet_name='FirstFile', index=False)
        dataFrameSecondFile.to_excel(writer, sheet_name='SecondFile', index=False)

    workbook = load_workbook(outputFilePath )
    differenceSheet = workbook.create_sheet(title='Differences')
    differenceSheet.append(['Column Name', 'Cell', 'FirstFile Value', 'SecondFile Value'])
    worksheetOne = workbook['FirstFile']
    worksheetTwo = workbook['SecondFile']

    for row in range(2, reshapedRows + 2):
        for col in range(1, reshapedColumns + 1):
            worksheetOneValue = worksheetOne.cell(row=row, column=col).value
            worksheetTwoValue = worksheetTwo.cell(row=row, column=col).value
            if worksheetOneValue != worksheetTwoValue:
                worksheetTwo.cell(row=row, column=col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                columnHeader = worksheetOne.cell(row=1, column=col).value if worksheetOne.cell(row=1, column=col).value else f"Col_{col}"
                cell_name = f"{chr(64 + col)}{row}" if col <= 26 else f"{chr(64 + (col - 1) // 26)}{chr(64 + col % 26)}{col}"
                differenceSheet.append([columnHeader, cell_name, worksheetOneValue, worksheetTwoValue])

    workbook.save(outputFilePath)
    print(f"Comparison complete! Differences highlighted and saved in: '{outputFilePath}'")
    print("\nExiting program...")


if __name__ == "__main__":
    compare_excel_files('File1.xlsx', 'File2.xlsx', 'Compared_Output.xlsx')
